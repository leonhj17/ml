# _*_ encoding:utf-8 _*_
from openpyxl import load_workbook
import numpy as np
import pandas as pd

pnum = 92
gnum = 9
ptnum = 7

filepath = 'bengbugz.xlsx'

def table_001(pnum, gnum, ptnum, sheet='001'):
    # 忽略管点，满足屏
    wb = load_workbook(filepath)
    sheet = wb[sheet]
    orig = np.zeros([pnum,1], dtype=(np.str, 8))

    # 原始数据，数据所在单元格，读入np.narray数组，数据格式 str8
    index = 0
    for row in sheet.rows:
        orig[index] = row[0].value
        index += 1

    # 根据屏号，管号，点号，生成空数据，并按规律写入原始数据
    res = np.zeros([pnum*gnum*ptnum, 1], dtype=(np.str, 8))
    for p in range(pnum):
        for g in range(gnum):
            for pt in range(ptnum):
                id = pt + g*ptnum + p*ptnum*gnum
                res[id] = orig[p]
    df = pd.DataFrame(res)
    return df


def table_002(pnum, gnum, ptnum, sheet='002'):
    # 忽略点，满足屏管
    wb = load_workbook(filepath)
    sheet = wb[sheet]
    orig = np.zeros([gnum, pnum], dtype=(np.str, 8))

    index = 0
    for row in sheet.rows:
        for col in range(pnum):
            orig[index,col] = row[col].value
        index += 1

    res = np.zeros([pnum*gnum*ptnum, 1], dtype=(np.str, 8))
    for p in range(pnum):
        for g in range(gnum):
            for pt in range(ptnum):
                id = pt + g*ptnum + p*ptnum*gnum
                res[id] = str(orig[g,p])
    df = pd.DataFrame(res)
    return df

def table_003(pnum, gnum, ptnum, sheet):
    # 忽略屏，满足管点
    wb = load_workbook(filepath)
    sheet = wb[sheet]
    orig = np.zeros([ptnum, gnum], dtype=(np.str, 8))

    index = 0
    for row in sheet.rows:
        for col in range(gnum):
            orig[index,col] = row[col].value
        index += 1

    res = np.zeros([pnum*gnum*ptnum, 1], dtype=(np.str, 8))
    for p in range(pnum):
        for g in range(gnum):
            for pt in range(ptnum):
                id = pt + g*ptnum + p*ptnum*gnum
                res[id] = str(orig[pt, g])
    df = pd.DataFrame(res)
    return df


def table_004(pnum, gnum, ptnum, sheet):
    # 忽略屏管，满足点
    wb = load_workbook(filepath)
    st = wb[sheet]
    orig = np.zeros([ptnum, 1], dtype=(np.str, 8))

    index = 0
    for row in st.rows:
        orig[index] = row[0].value
        index += 1

    res = np.zeros([pnum * gnum * ptnum, 1], dtype=(np.str, 8))
    for p in range(pnum):
        for g in range(gnum):
            for pt in range(ptnum):
                id = pt + g * ptnum + p * ptnum * gnum
                res[id] = orig[pt]
    df = pd.DataFrame(res)
    # df.to_excel('result.xlsx', sheet_name=sheet)
    return df

if __name__ =="__main__":
    df1 = table_001(pnum, gnum, ptnum)
    df2 = table_002(pnum, gnum, ptnum)
    df3 = table_003(pnum, gnum, ptnum, sheet='003')
    df4 = table_003(pnum, gnum, ptnum, sheet='004')
    df5 = table_003(pnum, gnum, ptnum, sheet='005')
    df6 = table_004(pnum, gnum, ptnum, sheet='006')
    df7 = table_004(pnum, gnum, ptnum, sheet='007')

    # 单次写入多个sheet
    with pd.ExcelWriter('result.xlsx') as writer:
        df1.to_excel(writer, sheet_name='001')
        df2.to_excel(writer, sheet_name='002')
        df3.to_excel(writer, sheet_name='003')
        df4.to_excel(writer, sheet_name='004')
        df5.to_excel(writer, sheet_name='005')
        df6.to_excel(writer, sheet_name='006')
        df7.to_excel(writer, sheet_name='007')